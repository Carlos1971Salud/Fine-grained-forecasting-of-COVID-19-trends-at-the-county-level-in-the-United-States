import numpy as np
import umap.umap_ as umap
import os
from time import sleep
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
from tqdm.auto import tqdm
from datetime import date
from scipy.signal import correlate
from scipy.io import loadmat, savemat
from sklearn.cluster import DBSCAN, SpectralClustering
from sklearn.metrics import silhouette_score

# The functions for clustering identification ==========================================================================

def Initial_UMAP_Distribution(data_file, save_path, data_ind_list):
    file_mat = loadmat(data_file, squeeze_me=True)
    data = file_mat['feature_dynamic']

    for list_of_n_neighbors in [10, 20, 30]:
        for list_of_min_dist in [0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9]:

            reducer = umap.UMAP(n_components=2, random_state=10086, n_neighbors=list_of_n_neighbors,
                                min_dist=list_of_min_dist, metric='correlation')
            reducer.fit(data)
            embedding = reducer.transform(data)

            if not os.path.exists(save_path + '{}_New'.format(data_ind_list)):
                os.makedirs(save_path + '{}_New'.format(data_ind_list))
                print("The new directory is created!")

            savepath = save_path + '{}_New/'.format(data_ind_list)
            filename = savepath + '3_component_umap_n_neighbors' + str(list_of_n_neighbors) + '_min_dist_' + str(
                list_of_min_dist) + '.npy'
            np.save(filename, embedding)

    print(data_ind_list + ' is finished for UMAP Processing=========')

    return savepath

def Initial_Clustering(Data_file, initial_savepath, save_path, list_of_n_neighbors, list_of_min_dist, max_cluster_num, data_ind_list):
    Initial_filename = initial_savepath + '3_component_umap_n_neighbors' + str(list_of_n_neighbors) + '_min_dist_' \
                       + str(list_of_min_dist) + '.npy'
    tempdata = np.load(Initial_filename)

    # Use the DBSCAN and Silhouette score to get the initial number of clusters
    choosing_cluster = []
    eps_list = np.arange(2, max_cluster_num)

    for eps_tmp in range(2, max_cluster_num):
        clustering = DBSCAN(eps=eps_tmp, min_samples=5).fit(tempdata)
        if max(clustering.labels_) == 0:
            choosing_cluster.append(0)
        else:
            choosing_cluster.append(silhouette_score(tempdata, clustering.labels_))

    eps_list_check = eps_list[np.where(choosing_cluster == max(choosing_cluster))]
    clustering_fin = DBSCAN(eps=eps_list_check[0], min_samples=5).fit(tempdata)

    filename_c1 = initial_savepath + '4_clustering_label_' + data_ind_list + '_1.npy'  # Temporally save the initial number of clusters
    filename_c2 = initial_savepath + '4_clustering_label_' + data_ind_list + '_2.npy'  # Temporally save the number of clusters we need

    initial_label = clustering_fin.labels_
    Num_cluster_need = max_cluster_num - max(initial_label)
    np.save(filename_c1, initial_label)
    np.save(filename_c2, Num_cluster_need)

    # Check the initial clusters being used to the subcluster processing
    if max(initial_label) < max_cluster_num:
        checked_list = []
        for j in range(max(initial_label)):
            checked_list.append(counting(initial_label, j))
        selected_cluster = checked_list.index(max(checked_list))

        filename_c3 = initial_savepath + '4_clustering_label_' + data_ind_list + '_3.npy'
        np.save(filename_c3, selected_cluster)
        generated_umap(Data_file, save_path, initial_label, selected_cluster, data_ind_list)
    else:
        print(data_ind_list + ' is finished for Initial Clustering=========')

    return tempdata, initial_label, Num_cluster_need

def generated_umap(Data_file, Save_Path, labels, selected_cluster, data_ind_list):
    file_mat = loadmat(Data_file, squeeze_me=True)
    data = file_mat['feature_dynamic']

    data = data[np.where(labels == selected_cluster), :]
    data = np.squeeze(data, axis=0)

    for list_of_n_neighbors_tmp in [10, 20, 30]:
        for list_of_min_dist_tmp in [0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9]:

            reducer = umap.UMAP(n_components=2, random_state=10086, n_neighbors=list_of_n_neighbors_tmp,
                                min_dist=list_of_min_dist_tmp, metric='correlation')  # correlation
            reducer.fit(data)
            embedding = reducer.transform(data)

            if not os.path.exists(Save_Path + '{}_New/Layer_1'.format(data_ind_list)):
                os.makedirs(Save_Path + '{}_New/Layer_1'.format(data_ind_list))
                print("The new directory is created!")

            savepath = Save_Path + '{}_New/Layer_1/'.format(data_ind_list)
            filename = savepath + '3_component_umap_n_neighbors' + str(list_of_n_neighbors_tmp) + '_min_dist_' + str(
                list_of_min_dist_tmp) + '.npy'
            np.save(filename, embedding)

    print(data_ind_list + ' is finished for Initial Clustering =========')

def counting(labels, ind_cluster):
    count_num = 0
    for ind in range(len(labels)):
        if labels[ind] == ind_cluster:
            count_num = count_num + 1
    return count_num

def subclustering_fun(save_path, initial_save_path, initial_label, Num_cluster_need, data_ind_list, list_of_min_dist):
    # Check to achieve the number of total clusters to the number we set
    if Num_cluster_need > 0:
        minimal_cluster_list = []
        parameter_list = []
        shi_score = []

        for list_of_n_neighbors_tmp in [10, 20]:

            savepath_1 = save_path + '{}_New/Layer_1/'.format(data_ind_list)
            filename = savepath_1 + '3_component_umap_n_neighbors' + str(list_of_n_neighbors_tmp) + '_min_dist_' + str(list_of_min_dist) + '.npy'
            sub_tempdata = np.load(filename)

            # Use the Spectral Clustering method to get the rest of subclusters
            clustering = SpectralClustering(n_clusters=(Num_cluster_need + 1)).fit(sub_tempdata)
            if len(minimal_cluster_list) == 0:
                minimal_cluster_list = clustering.labels_
            else:
                minimal_cluster_list = np.vstack((minimal_cluster_list, clustering.labels_))
            parameter_list.append(list_of_n_neighbors_tmp)
            shi_score.append(silhouette_score(sub_tempdata, clustering.labels_))

        index, = np.where(shi_score == max(shi_score))

        current_labels = minimal_cluster_list[np.squeeze(index[0]), :]
        selected_parameters = parameter_list[np.squeeze(index[0])]
        filename_c3 = initial_save_path + '5_subclustering_label_' + data_ind_list + '.npy'  # Save the subcluster label
        filename_c4 = initial_save_path + '5_subclustering_label_Extra_' + data_ind_list + '.npy'  # Save the parameters we got
        np.save(filename_c3, current_labels)
        np.save(filename_c4, selected_parameters)

        # reorder list==============================================================================================
        filename_c5 = initial_save_path + '4_clustering_label_' + data_ind_list + '_3.npy'
        selected_cluster = np.load(filename_c5)
        subcluster_num = np.unique(current_labels)

        tmp_max = max(initial_label)
        current_labels_New = np.zeros(current_labels.shape, dtype=int)
        for pos in range(len(subcluster_num)):
            if pos == 0:
                current_labels_New[np.where(current_labels == subcluster_num[pos])] = selected_cluster
            else:
                current_labels_New[np.where(current_labels == subcluster_num[pos])] = tmp_max + pos

        temp_final_label = initial_label
        temp_final_label[np.where(initial_label == selected_cluster)] = current_labels_New

    else:
        temp_final_label = initial_label

    print(np.unique(temp_final_label))

    return temp_final_label

def Reordering_Clusters(ori_Data_file, initial_save_path, temp_final_label, list_of_n_neighbors, list_of_min_dist, data_ind_list):

    Final_label = np.zeros((len(temp_final_label),), dtype=int)
    all_mat = loadmat(ori_Data_file)
    Infection_days = all_mat['nop_increase_total_trimmed']
    mean_clustering_infection = []

    for m_c in np.unique(temp_final_label):
        temp_mean = Infection_days[np.where(temp_final_label == m_c)]
        mean_clustering_infection.append(np.mean(temp_mean[:, (temp_mean.shape[1]) - 7:]))

    mean_indices = np.argsort(mean_clustering_infection)
    mean_indices = mean_indices[::-1]

    for ind_c in range(len(mean_indices)):
        Final_label[np.where(temp_final_label == mean_indices[ind_c])] = (ind_c + 1)

    filename_all = initial_save_path + '3_component_umap_n_neighbors' + str(list_of_n_neighbors) + '_min_dist_' + str(
        list_of_min_dist) + '.npy'
    tempdata = np.load(filename_all)

    # Save the final clustering list and the UMAP distribution
    savemat(initial_save_path + '6_Final_clustering_label_{}.mat'.format(data_ind_list),
            mdict={'Final_label': Final_label, 'Umap_pos': tempdata})

    # if index % 10 == 0:
    #     plt.figure()
    #     plt.scatter(tempdata[:, 0], tempdata[:, 1], c=Final_label)
    #     plt.legend()

    return Final_label

def plot_series(time, series, format='-', start=0, end=None):
    plt.plot(time[start:end], series[start:end], format)
    plt.xlabel("Date")
    plt.ylabel("Number")
    plt.grid(True)

def train_window_ungraphing(df, ref_day, predict_day):
    X_train, Y_train = [], []
    for i in range(df.shape[0]-predict_day-ref_day):
        X_train.append(np.array(df[i:i+ref_day]))
        Y_train.append(np.array(df[i+ref_day: i+ref_day+predict_day]))
    return np.array(X_train), np.array(Y_train)

def shuffle(X, Y, shuffle_buffer):
    np.random.seed(shuffle_buffer)
    randmList = np.arange(X.shape[0])
    np.random.shuffle(randmList)
    return X[randmList], Y[randmList]

# The functions for evaluation =========================================================================================

def RMSE_Measure(truth, Pred):
    difference_array = np.subtract(truth, Pred)
    squared_array = np.square(difference_array)
    row_mean_array = np.mean(squared_array, axis=1)
    sqrt_array = np.sqrt(row_mean_array)

    return sqrt_array

def MSE_Measure(truth, Pred):
    difference_array = np.subtract(truth, Pred)
    squared_array = np.square(difference_array)
    row_mean_array = np.mean(squared_array, axis=1)

    return row_mean_array

def MAPE_Measure(truth, Pred):
    difference_array = np.subtract(truth, Pred)
    truth[np.where(truth == 0)] = 1
    different_truth = (difference_array / truth)
    squared_array = np.abs(different_truth)
    row_mean_array = np.mean(squared_array, axis=1)

    return row_mean_array*100

def RRMSE_Measure_1(truth, Pred):
    difference_array = np.subtract(truth, Pred)
    squared_array = np.square(difference_array)
    row_mean_array = np.mean(squared_array, axis=1)
    sqrt_array = np.sqrt(row_mean_array)

    truth_squared_array = np.square(truth)
    truth_mean_array = np.mean(truth_squared_array, axis=1)
    truth_sqrt_array = np.sqrt(truth_mean_array)

    rrmse_array = sqrt_array/truth_sqrt_array

    return rrmse_array*100

def rolling_value(seq_array, window):

    new_array = np.zeros((seq_array.shape[0], seq_array.shape[1]))
    for i in range(seq_array.shape[1]-window):
        new_array[:, i+window] = np.round(np.mean(seq_array[:, i+window - 7: i+window], axis=1))
    for j in range(window):
        if j > 0:
            new_array[:, j] = np.round(np.mean(seq_array[:, 0:j+1], axis=1))

    return new_array

def save_to_file(save_path, time_list, data):
    df_all_cases = pd.read_excel(save_path, sheet_name='Cases')
    df_all_cases_County = df_all_cases['County']
    df_all_cases_State = df_all_cases['State']
    workbook = openpyxl.load_workbook(save_path)
    if not 'Clusters1' in workbook.sheetnames:
        workbook.create_sheet('Clusters1')
    worksheet = workbook['Clusters1']
    for pos, val in enumerate(df_all_cases_County):
        if pos == 0:
            worksheet.cell(row=pos + 1, column=1).value = 'County'
            worksheet.cell(row=pos + 1, column=2).value = 'State'
        worksheet.cell(row=pos + 2, column=1).value = val
        worksheet.cell(row=pos + 2, column=2).value = df_all_cases_State[pos]
        for lab_col in range(len(time_list)):
            if pos == 0:
                worksheet.cell(row=1, column=lab_col + 3).value = time_list[lab_col]
            worksheet.cell(row=pos + 2, column=lab_col + 3).value = data[pos, lab_col]
    workbook.save(save_path)
    workbook.close()

def autocorrelation_function(x):
    correlation = correlate(x, x, mode='full', method='fft')
    autocorr = correlation[len(x) - 1:] / np.max(correlation)

    return autocorr

# the function for cleaning prediction results =========================================================================
def clean_prediction_data(length_day, assign_label, tmp_predicted_list):
    index1, = np.where(assign_label > 1)
    index2, = np.where(assign_label == 0)

    if len(index1) != 0:
        for idx in index1:
            tmp_predicted_list[:, idx] = tmp_predicted_list[:, idx] / assign_label[idx]

    if len(index2) != 0:
        for idx in index2:
            if idx == 0:
                tmp_predicted_list[:, idx] = (tmp_predicted_list[:, idx + 1] + tmp_predicted_list[:, idx + 2]) / 2
            elif idx == (length_day.days - 1):
                tmp_predicted_list[:, idx] = (tmp_predicted_list[:, idx - 1] + tmp_predicted_list[:, idx - 2]) / 2
            else:
                tmp_predicted_list[:, idx] = (tmp_predicted_list[:, idx - 1] + tmp_predicted_list[:, idx + 1]) / 2

    tmp_predicted_list[tmp_predicted_list < 0] = 0

    return tmp_predicted_list

# main function for combining prediction results (for 15 day-length)

def Post_processing(save_path, Prediction_path, time_list, All_Time_Period, df_New_cases_diff_denoised):

    pre_pix = 7
    post_pix = -17

    start_date = date(2020, 1, 21)
    actual_start_date = date(2020, 4, 15)
    Final_date = date(2022, 4, 15)
    x_length = actual_start_date - start_date
    Time_length = Final_date - start_date
    length_day = Final_date - actual_start_date
    Period_time_start = Time_length.days - length_day.days
    Period_time_end = Time_length.days

    truth_day_report = np.mean(df_New_cases_diff_denoised, axis=0)
    assign_label = np.zeros(length_day.days)

    for c_ind in tqdm(range(len(time_list))):

        file_path = Prediction_path + '{}_Result_Save/'.format(time_list[c_ind])

        tmp_time = All_Time_Period[c_ind]
        if len(tmp_time) == 8:
            Predict_Start_date = date((2000 + int(tmp_time[6:8])), int(tmp_time[0:2]), int(tmp_time[3:5]))
        else:
            Predict_Start_date = date((2000 + int(tmp_time[5:7])), int(tmp_time[0:1]), int(tmp_time[2:4]))

        Time_length = Predict_Start_date - actual_start_date
        list_post = Time_length.days + 15  # 15 day-length
        list_pre = Time_length.days

        record_list = np.zeros(df_New_cases_diff_denoised.shape[0])
        arr = os.listdir(file_path)

        tmp_predicted_trend_1 = []
        tmp_predicted_trend_2 = []
        tmp_predicted_trend_3 = []
        tmp_predicted_trend_4 = []
        tmp_predicted_trend_5 = []
        tmp_predicted_trend_6 = []
        tmp_predicted_trend_7 = []
        tmp_predicted_trend_8 = []
        tmp_predicted_trend_9 = []
        tmp_predicted_trend_10 = []
        tmp_predicted_trend_11 = []
        tmp_predicted_trend_12 = []
        tmp_predicted_trend_13 = []
        tmp_predicted_trend_14 = []
        tmp_predicted_trend_15 = []

        for a_c_i in range(df_New_cases_diff_denoised.shape[0]):
            record_list[a_c_i] = record_list[a_c_i] + 1
            for d_i in arr:
                if d_i[-7] == '_':
                    post_pix_1 = post_pix - 1
                else:
                    post_pix_1 = post_pix
                if str(a_c_i) == d_i[pre_pix:post_pix_1]:
                    tmp_data = loadmat(file_path + d_i)
                    tmp_all = tmp_data['Sliding_predicted_list']

                    if len(tmp_predicted_trend_1) == 0:
                        tmp_predicted_trend_1 = tmp_all[0, :]
                        tmp_predicted_trend_2 = tmp_all[1, :]
                        tmp_predicted_trend_3 = tmp_all[2, :]
                        tmp_predicted_trend_4 = tmp_all[3, :]
                        tmp_predicted_trend_5 = tmp_all[4, :]
                        tmp_predicted_trend_6 = tmp_all[5, :]
                        tmp_predicted_trend_7 = tmp_all[6, :]
                        tmp_predicted_trend_8 = tmp_all[7, :]
                        tmp_predicted_trend_9 = tmp_all[8, :]
                        tmp_predicted_trend_10 = tmp_all[9, :]
                        tmp_predicted_trend_11 = tmp_all[10, :]
                        tmp_predicted_trend_12 = tmp_all[11, :]
                        tmp_predicted_trend_13 = tmp_all[12, :]
                        tmp_predicted_trend_14 = tmp_all[13, :]
                        tmp_predicted_trend_15 = tmp_all[14, :]
                    else:
                        tmp_predicted_trend_1 = np.vstack((tmp_predicted_trend_1, tmp_all[0, :]))
                        tmp_predicted_trend_2 = np.vstack((tmp_predicted_trend_2, tmp_all[1, :]))
                        tmp_predicted_trend_3 = np.vstack((tmp_predicted_trend_3, tmp_all[2, :]))
                        tmp_predicted_trend_4 = np.vstack((tmp_predicted_trend_4, tmp_all[3, :]))
                        tmp_predicted_trend_5 = np.vstack((tmp_predicted_trend_5, tmp_all[4, :]))
                        tmp_predicted_trend_6 = np.vstack((tmp_predicted_trend_6, tmp_all[5, :]))
                        tmp_predicted_trend_7 = np.vstack((tmp_predicted_trend_7, tmp_all[6, :]))
                        tmp_predicted_trend_8 = np.vstack((tmp_predicted_trend_8, tmp_all[7, :]))
                        tmp_predicted_trend_9 = np.vstack((tmp_predicted_trend_9, tmp_all[8, :]))
                        tmp_predicted_trend_10 = np.vstack((tmp_predicted_trend_10, tmp_all[9, :]))
                        tmp_predicted_trend_11 = np.vstack((tmp_predicted_trend_11, tmp_all[10, :]))
                        tmp_predicted_trend_12 = np.vstack((tmp_predicted_trend_12, tmp_all[11, :]))
                        tmp_predicted_trend_13 = np.vstack((tmp_predicted_trend_13, tmp_all[12, :]))
                        tmp_predicted_trend_14 = np.vstack((tmp_predicted_trend_14, tmp_all[13, :]))
                        tmp_predicted_trend_15 = np.vstack((tmp_predicted_trend_15, tmp_all[14, :]))

        # ====================================================================================================
        if c_ind == 0:
            tmp_predicted_list_1 = np.zeros((df_New_cases_diff_denoised.shape[0], length_day.days))
            tmp_predicted_list_2 = np.zeros((df_New_cases_diff_denoised.shape[0], length_day.days))
            tmp_predicted_list_3 = np.zeros((df_New_cases_diff_denoised.shape[0], length_day.days))
            tmp_predicted_list_4 = np.zeros((df_New_cases_diff_denoised.shape[0], length_day.days))
            tmp_predicted_list_5 = np.zeros((df_New_cases_diff_denoised.shape[0], length_day.days))
            tmp_predicted_list_6 = np.zeros((df_New_cases_diff_denoised.shape[0], length_day.days))
            tmp_predicted_list_7 = np.zeros((df_New_cases_diff_denoised.shape[0], length_day.days))
            tmp_predicted_list_8 = np.zeros((df_New_cases_diff_denoised.shape[0], length_day.days))
            tmp_predicted_list_9 = np.zeros((df_New_cases_diff_denoised.shape[0], length_day.days))
            tmp_predicted_list_10 = np.zeros((df_New_cases_diff_denoised.shape[0], length_day.days))
            tmp_predicted_list_11 = np.zeros((df_New_cases_diff_denoised.shape[0], length_day.days))
            tmp_predicted_list_12 = np.zeros((df_New_cases_diff_denoised.shape[0], length_day.days))
            tmp_predicted_list_13 = np.zeros((df_New_cases_diff_denoised.shape[0], length_day.days))
            tmp_predicted_list_14 = np.zeros((df_New_cases_diff_denoised.shape[0], length_day.days))
            tmp_predicted_list_15 = np.zeros((df_New_cases_diff_denoised.shape[0], length_day.days))

        tmp_predicted_list_1[:, list_pre: list_post] = tmp_predicted_list_1[:, list_pre: list_post] + tmp_predicted_trend_1
        tmp_predicted_list_2[:, list_pre: list_post] = tmp_predicted_list_2[:, list_pre: list_post] + tmp_predicted_trend_2
        tmp_predicted_list_3[:, list_pre: list_post] = tmp_predicted_list_3[:, list_pre: list_post] + tmp_predicted_trend_3
        tmp_predicted_list_4[:, list_pre: list_post] = tmp_predicted_list_4[:, list_pre: list_post] + tmp_predicted_trend_4
        tmp_predicted_list_5[:, list_pre: list_post] = tmp_predicted_list_5[:, list_pre: list_post] + tmp_predicted_trend_5
        tmp_predicted_list_6[:, list_pre: list_post] = tmp_predicted_list_6[:, list_pre: list_post] + tmp_predicted_trend_6
        tmp_predicted_list_7[:, list_pre: list_post] = tmp_predicted_list_7[:, list_pre: list_post] + tmp_predicted_trend_7
        tmp_predicted_list_8[:, list_pre: list_post] = tmp_predicted_list_8[:, list_pre: list_post] + tmp_predicted_trend_8
        tmp_predicted_list_9[:, list_pre: list_post] = tmp_predicted_list_9[:, list_pre: list_post] + tmp_predicted_trend_9
        tmp_predicted_list_10[:, list_pre: list_post] = tmp_predicted_list_10[:, list_pre: list_post] + tmp_predicted_trend_10
        tmp_predicted_list_11[:, list_pre: list_post] = tmp_predicted_list_11[:, list_pre: list_post] + tmp_predicted_trend_11
        tmp_predicted_list_12[:, list_pre: list_post] = tmp_predicted_list_12[:, list_pre: list_post] + tmp_predicted_trend_12
        tmp_predicted_list_13[:, list_pre: list_post] = tmp_predicted_list_13[:, list_pre: list_post] + tmp_predicted_trend_13
        tmp_predicted_list_14[:, list_pre: list_post] = tmp_predicted_list_14[:, list_pre: list_post] + tmp_predicted_trend_14
        tmp_predicted_list_15[:, list_pre: list_post] = tmp_predicted_list_15[:, list_pre: list_post] + tmp_predicted_trend_15

        assign_label[list_pre: list_post] = assign_label[list_pre: list_post] + 1
    sleep(0.001)

    all_sliding_day_1 = clean_prediction_data(length_day, assign_label, tmp_predicted_list_15)
    all_sliding_day_2 = clean_prediction_data(length_day, assign_label, tmp_predicted_list_14)
    all_sliding_day_3 = clean_prediction_data(length_day, assign_label, tmp_predicted_list_13)
    all_sliding_day_4 = clean_prediction_data(length_day, assign_label, tmp_predicted_list_12)
    all_sliding_day_5 = clean_prediction_data(length_day, assign_label, tmp_predicted_list_11)
    all_sliding_day_6 = clean_prediction_data(length_day, assign_label, tmp_predicted_list_10)
    all_sliding_day_7 = clean_prediction_data(length_day, assign_label, tmp_predicted_list_9)
    all_sliding_day_8 = clean_prediction_data(length_day, assign_label, tmp_predicted_list_8)
    all_sliding_day_9 = clean_prediction_data(length_day, assign_label, tmp_predicted_list_7)
    all_sliding_day_10 = clean_prediction_data(length_day, assign_label, tmp_predicted_list_6)
    all_sliding_day_11 = clean_prediction_data(length_day, assign_label, tmp_predicted_list_5)
    all_sliding_day_12 = clean_prediction_data(length_day, assign_label, tmp_predicted_list_4)
    all_sliding_day_13 = clean_prediction_data(length_day, assign_label, tmp_predicted_list_3)
    all_sliding_day_14 = clean_prediction_data(length_day, assign_label, tmp_predicted_list_2)
    all_sliding_day_15 = clean_prediction_data(length_day, assign_label, tmp_predicted_list_1)

    # Persistence Model Prediction
    truth_day_report_act = df_New_cases_diff_denoised[:, x_length.days - 1:-3]
    truth_day_report_1_before = df_New_cases_diff_denoised[:, x_length.days - 2:-4]
    truth_day_report_2_before = df_New_cases_diff_denoised[:, x_length.days - 3:-5]
    truth_day_report_3_before = df_New_cases_diff_denoised[:, x_length.days - 4:-6]
    truth_day_report_4_before = df_New_cases_diff_denoised[:, x_length.days - 5:-7]
    truth_day_report_5_before = df_New_cases_diff_denoised[:, x_length.days - 6:-8]
    truth_day_report_6_before = df_New_cases_diff_denoised[:, x_length.days - 7:-9]
    truth_day_report_7_before = df_New_cases_diff_denoised[:, x_length.days - 8:-10]
    truth_day_report_8_before = df_New_cases_diff_denoised[:, x_length.days - 9:-11]
    truth_day_report_9_before = df_New_cases_diff_denoised[:, x_length.days - 10:-12]
    truth_day_report_10_before = df_New_cases_diff_denoised[:, x_length.days - 11:-13]
    truth_day_report_11_before = df_New_cases_diff_denoised[:, x_length.days - 12:-14]
    truth_day_report_12_before = df_New_cases_diff_denoised[:, x_length.days - 13:-15]
    truth_day_report_13_before = df_New_cases_diff_denoised[:, x_length.days - 14:-16]
    truth_day_report_14_before = df_New_cases_diff_denoised[:, x_length.days - 15:-17]
    truth_day_report_15_before = df_New_cases_diff_denoised[:, x_length.days - 16:-18]

    if not os.path.exists(save_path):
        os.makedirs(save_path)

    # Save the prediction results
    savemat(save_path + 'Infection_prediction_record.mat',
            mdict={'truth_list': truth_day_report_act, 'Prediction_list': all_sliding_day_1,
                   'Prediction_list_2_days': all_sliding_day_2, 'Prediction_list_3_days': all_sliding_day_3,
                   'Prediction_list_4_days': all_sliding_day_4, 'Prediction_list_5_days': all_sliding_day_5,
                   'Prediction_list_6_days': all_sliding_day_6, 'Prediction_list_7_days': all_sliding_day_7,
                   'Prediction_list_8_days': all_sliding_day_8, 'Prediction_list_9_days': all_sliding_day_9,
                   'Prediction_list_10_days': all_sliding_day_10, 'Prediction_list_11_days': all_sliding_day_11,
                   'Prediction_list_12_days': all_sliding_day_12, 'Prediction_list_13_days': all_sliding_day_13,
                   'Prediction_list_14_days': all_sliding_day_14, 'Prediction_list_15_days': all_sliding_day_15,
                   'Persistence_1': truth_day_report_1_before, 'Persistence_7': truth_day_report_7_before})

    print('All Forecasting Results are completely obtained !!! ============================================================')
